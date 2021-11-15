from oauth2client.client import verify_id_token
from openpyxl import load_workbook # 파일 불러오기
import os

def po_bl_num_so():
    wb = load_workbook("Walmart PIckups GA Warehouse.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
    ws = wb.active  # 활성화된 Sheet

    Lod_num = []  # 2-1-1-1
    Lod_num_bl = [] # 2-1-1-2


    for x in range(2, 150):# 2-1-1-3
        y=4 # 2-1-1-3-1
        verify=ws.cell(row=x, column=y).value # 2-1-1-3-2
        if verify == None: #
            # 2-1-1-3-3-1
            #Lod_num_bl.append(ws.cell(row=x, column=y)) 
            Lod_num_bl.append([x,y]) # << Nested Lists(중첩 리스트 생성 )


        else: # 2-1-1-3-4
            #Lod_num.append(ws.cell(row=x, column=y)) 
            Lod_num.append([x,y]) # 2-1-1-3-4-1


    # List index 값확인
    # index = Lod_num_bl.index([9,4]) 
    # print('The index of 9,4:', index)

    PO_num_li = []

    for y in range(len(Lod_num_bl)): # 2-1-2
        Lod_num_bl[y][1]-=2 # 2-1-2-1
        
    for z in range(len(Lod_num_bl)): # 2-1-2-3
        for k in range(0,1):
            PO_num=ws.cell(row=Lod_num_bl[z][k], column=Lod_num_bl[z][1]).value # 2-1-2-3-1
            # print("row:"+str(Lod_num_bl[z][k]))
            # print("column:"+str(Lod_num_bl[z][1]))
            # print("PO_num_li:"+str(PO_num_li))
            PO_num_li.append(PO_num) # 2-1-2-3-2

    #print(Lod_num_bl) # None 인 모든 셀 정보 출력
    #print(Lod_num) # None이 아닌 모든 셀 정보 출력
    print("Load 넘버가 없는 모든 PO번호:"+ str(PO_num_li)) # Load 넘버가 없는 모든 PO번호 출력





def insert_load_num():
    ## 현재 폴더에서 파일 이름 리스트를 얻기
    path = "C:/Users/zinus_user/Downloads"
    file_list = os.listdir(path)
    print ("file_list: {}".format(file_list))
    # file_list: ['test2.xlsx', 'testdir', 'test1.xlsx']

    ## 파일중 엑셀파일 리스트 얻기
    exfile_list =[i for i in file_list if os.path.splitext(i)[-1]==".xlsx"]
    print("excel:", exfile_list)
    # excel: ['test2.xlsx', 'test1.xlsx']
    
    excelfilename = exfile_list[0]
    wb = load_workbook(str(path+"/"+excelfilename)) # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
    ws = wb.active  # 활성화된 Sheet

    ## 엑셀 시트 이름 얻기
    # wb = load_workbook(filename = "/Users/xzero/excel_test/test1.xlsx")
    # print (wb.sheetnames)
    # ['1번', 'two']

    load_num1 = ws.cell(row=2, column=2).value 
    load_num2 = ws.cell(row=2, column=9).value 
    load_num3 = ws.cell(row=2, column=16).value 
    print(load_num1, load_num2, load_num3)
    
    wb = load_workbook("Walmart PIckups GA Warehouse.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
    ws = wb.active  # 활성화된 Sheet




po_bl_num_so()
insert_load_num()




