import pyautogui 
import pyperclip
import time
from openpyxl import load_workbook
from random import *
import schedule
from datetime import datetime
import os
# def start_up():
#     pyautogui.moveTo(27, 1052) # 1-2
#     time.sleep(2)
#     pyautogui.click() # 1-3
#     time.sleep(2)
#     pyperclip.copy("SAP") # 1-4
#     pyautogui.hotkey('ctrl', 'v') # 1-5
#     time.sleep(2)
#     pyautogui.press('enter') # 1-6
#     time.sleep(2)
#     pyautogui.press('enter') # 1-7
#     time.sleep(2)
#     pyautogui.moveTo(289, 279) # 1-8
#     time.sleep(2)
#     pyautogui.click() # 1-9
#     time.sleep(2)
#     pyperclip.copy("E00145") # 1-10 E00127
#     pyautogui.hotkey('ctrl', 'v') # 1-11
#     pyautogui.moveTo(280, 320) # 1-12
#     pyautogui.click() # 1-13
#     pyperclip.copy("Cjl0125!") # 1-14 Cjl011722!
#     pyautogui.hotkey('ctrl', 'v') # 1-15
#     time.sleep(2)
#     pyautogui.press('enter') # 1-16
#     time.sleep(2)
#     pyautogui.moveTo(118, 129) # 1-17
#     pyautogui.click() # 1-18
#     pyperclip.copy("ZWMR0224") # 1-19
#     pyautogui.hotkey('ctrl', 'v') # 1-20
#     pyautogui.press('enter') # 1-21

def empty_bin_macro_start():
    try:
        print("Started SAP Macro system")
        print("start operation in 5 sec")

        k=1
        for k in range(5):
            print(k)
            time.sleep(5)

        print("system start!")

        wb = load_workbook(os.path.join(os.getcwd(), "SBins.xlsx")) # 2-24
        ws = wb.active  # 2-25

        x=1
        y=1
        while True:
            verify = ws.cell(row=x, column=y).value # 2-24

            if verify != None:
                # s=datetime.now().strftime('%H')
                # if int(s) >=11 and int(s) < 12 :
                #     time.sleep(3699)
                # if int(s) >= 15 and int(s) < 16:
                #     time.sleep(3699)
                # if int(s) >= 20:
                #     break
                
                pyautogui.moveTo(195, 360) # 1-22
                pyautogui.click() # 1-23
                
                pyautogui.press('enter') # 1-27
                pyautogui.press('delete') # 1-32
                
                i = randint(3, 7)
                print(str(verify))
                time.sleep(2)
                pyperclip.copy(verify) # 1-25
                pyautogui.hotkey('ctrl', 'v') # 1-26
                pyautogui.press('enter') # 1-27
                time.sleep(2)

                pyautogui.moveTo(211, 198) # 1-28
                pyautogui.click() # 1-29
                time.sleep(i)

                pyautogui.moveTo(139, 356) # 1-30
                pyautogui.click() # 1-31
                time.sleep(3)
                x+=1
                
            else:
                time.sleep(2)
                pyautogui.moveTo(1886, 16) # 1-33
                pyautogui.click() # 1-34
                time.sleep(2)
                pyautogui.moveTo(350, 614) # 1-35
                pyautogui.click() # 1-36
                time.sleep(2)
                print("All task has been done. Last value:: ["+str(verify)+"]")
                print("Please, restart program with new Data sheet")
                break

    except Exception as e :
        print("Error detail: [ "+e+" ]")
        print("Please, restart system again")

# start_up()
empty_bin_macro_start()
# schedule.every().monday.at("12:44").do(empty_bin_macro_start)
# schedule.every().saturday.at("08:00").do(empty_bin_macro_start)
# schedule.every().sunday.at("08:00").do(empty_bin_macro_start)

