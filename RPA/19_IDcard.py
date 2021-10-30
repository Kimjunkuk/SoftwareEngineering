from openpyxl import load_workbook # openpyxl 모듈의 load_workbook 함수 불러오기
from openpyxl.drawing.image import Image # openpyxl 모듈의 Image 함수 불러오기

wb = load_workbook("EmployeeIDcard.xlsx") # 작업을 원하는 워크북 불러오기
ws = wb.active # 현제 활성화된 sheet 가져옴

img = Image("CJL-new logo-14.png") # img 변수에 이미지 파일을 저장

# 이미지 사이즈 조절 하기
img.height = 72.96
img.width = 212.16


wb["ID Cards"].add_image.cell(img, "A2") # 워크북의 ID Cards 시트의 A4 셀에 이미지 삽입 

# print(wb.sheetnames) # 모든 Sheet 이름 확인
# print(wb.sheetnames)


# Dict 형태로 Sheet 에 접근
new_ws1 = wb["Sheet2"]  
new_ws2 = wb["ID Cards"] 

# target 변수에 Sheet2 에 A2 셀에 있는 저장
target1 = new_ws1["A2"].value

# ID Cards 시트에 있는 A1셀에 붙여 넣기 
new_ws2["A1"] = target1


wb.save("EmployeeIDcard.xlsx")
