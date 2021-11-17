from openpyxl import load_workbook # 파일 불러오기


PO_num_li = [] # Load 번호가 없는 PO 번호 전역 리스트


def po_bl_num_so():
    wb = load_workbook("Walmart PIckups GA Warehouse.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
    ws = wb.active  # 활성화된 Sheet

    Lod_num = []  # 2-1-1-1
    Lod_num_bl = [] # 2-1-1-2


    for x in range(2, 150):# 2-1-1-3
        y=4 # 2-1-1-3-1
        verify=ws.cell(row=x, column=y).value # 2-1-1-3-2
        if verify == None: #
            # 2-1-1-3-3-1
            #Lod_num_bl.append(ws.cell(row=x, column=y)) 
            Lod_num_bl.append([x,y]) # << Nested Lists(중첩 리스트 생성 )


        else: # 2-1-1-3-4
            #Lod_num.append(ws.cell(row=x, column=y)) 
            Lod_num.append([x,y]) # 2-1-1-3-4-1


    # List index 값확인
    # index = Lod_num_bl.index([9,4]) 
    # print('The index of 9,4:', index)


    for y in range(len(Lod_num_bl)): # 2-1-2
        Lod_num_bl[y][1]-=2 # 2-1-2-1
        
    for z in range(len(Lod_num_bl)): # 2-1-2-3
        for k in range(0,1):
            PO_num=ws.cell(row=Lod_num_bl[z][k], column=Lod_num_bl[z][1]).value # 2-1-2-3-1
            # print("row:"+str(Lod_num_bl[z][k]))
            # print("column:"+str(Lod_num_bl[z][1]))
            # print("PO_num_li:"+str(PO_num_li))
            PO_num_li.append(PO_num) # 2-1-2-3-2

    #print(Lod_num_bl) # None 인 모든 셀 정보 출력
    #print(Lod_num) # None이 아닌 모든 셀 정보 출력
    print("Load 넘버가 없는 모든 PO번호:"+ str(PO_num_li)) # Load 넘버가 없는 모든 PO번호 출력

        
    # load_num1 = ws.cell(row=2, column=2).value 
    # load_num2 = ws.cell(row=2, column=9).value 
    # load_num3 = ws.cell(row=2, column=16).value 
    # print(load_num1, load_num2, load_num3)
    
    # wb = load_workbook("Walmart PIckups GA Warehouse.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
    # ws = wb.active  # 활성화된 Sheet
    
#     print(ws.cell(column=1, row=1).value) # ws["A1"].value
# print(ws.cell(column=2, row=1).value) # ws["B1"].value

# c = ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10
# print(c.value) # ws["C1"]

# from random import *

# # 반복문을 이용해서 랜덤 숫자 채우기
# index = 1
# for x in range(1, 11): # 10 개 row
#     for y in range(1, 11): # 10 개 column
#         #ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자
#         ws.cell(row=x, column=y, value=index)
#         index += 1

# wb.save("sample.xlsx")


po_bl_num_so()
