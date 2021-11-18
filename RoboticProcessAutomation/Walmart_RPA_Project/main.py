from openpyxl import load_workbook # 파일 불러오기
from selenium import webdriver 
from selenium.webdriver.common.keys import Keys 
import time # 11/17 프로세스 흐름 제어, 추후 시간 제어 방식이 아닌 프로세스 직접 제어 변경 예정
import pyperclip 
import pyautogui
import os

from Google import Create_Service # download source from the link in the description
import pandas as pd

def run_batchUpdate_request(service, google_sheet_id, request_body_json):
    try:
        response = service.spreadsheets().batchUpdate(
            spreadsheetId=google_sheet_id,
            body=request_body_json
        ).execute()
        return response
    except Exception as e:
        print(e)
        return None

import os
from Google import Create_Service

CLIENT_SECRET_FILE = 'client_secret.json'
API_SERVICE_NAME = 'sheets'
API_VERSION = 'v4'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
GOOGLE_SHEET_ID = '1dnQiI-SDbUPXBCfKttT-ROx8nUKSmA7P_ExlNGXYALM'

service = Create_Service(CLIENT_SECRET_FILE, API_SERVICE_NAME, API_VERSION, SCOPES)

"""
Iterate Worksheets
"""
gsheets = service.spreadsheets().get(spreadsheetId=GOOGLE_SHEET_ID).execute()
sheets = gsheets['sheets']

for sheet in sheets:
    if sheet['properties']['title'] != 'master':
        dataset = service.spreadsheets().values().get(
            spreadsheetId=GOOGLE_SHEET_ID,
            range=sheet['properties']['title'],
            majorDimension='ROWS'
        ).execute()
        df = pd.DataFrame(dataset['values'])
        df.columns = df.iloc[0]
        df.drop(df.index[0], inplace=True)
        df.to_csv(sheet['properties']['title'] + '.csv', index=False)

import pandas as pd     # pip install pands필요
import numpy as np

# 해당이름의 csv파일을 읽어옴
r_csv = pd.read_csv("Sheet1.csv")

# 저장할 xlsx파일의 이름을 정함
save_xlsx = pd.ExcelWriter("Sheet1.xlsx")

r_csv.to_excel(save_xlsx, index = False) # xlsx 파일로 변환

save_xlsx.save() #xlsx 파일로 저장


PO_num_li = [] # Load 번호가 없는 PO 번호 전역 리스트

wb = load_workbook("Sheet1.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
ws = wb.active  # 활성화된 Sheet
ws.column_dimensions['B'].width= 50
wb.save("Sheet1.xlsx")

Lod_num = []  # 2-1-1-1
Lod_num_bl = [] # 2-1-1-2

for x in range(2, 150):# 2-1-1-3
    y=4 # 2-1-1-3-1
    verify=ws.cell(row=x, column=y).value # 2-1-1-3-2
    if verify == None: #
        # 2-1-1-3-3-1
        #Lod_num_bl.append(ws.cell(row=x, column=y)) 
        Lod_num_bl.append([x,y]) # << Nested Lists(중첩 리스트 생성 )


    else: # 2-1-1-3-4
        #Lod_num.append(ws.cell(row=x, column=y)) 
        Lod_num.append([x,y]) # 2-1-1-3-4-1


# List index 값확인
# index = Lod_num_bl.index([9,4]) 
# print('The index of 9,4:', index)


for y in range(len(Lod_num_bl)): # 2-1-2
    Lod_num_bl[y][1]-=2 # 2-1-2-1
    
for z in range(len(Lod_num_bl)): # 2-1-2-3
    for k in range(0,1):
        PO_num=ws.cell(row=Lod_num_bl[z][k], column=Lod_num_bl[z][1]).value # 2-1-2-3-1
        # print("row:"+str(Lod_num_bl[z][k]))
        # print("column:"+str(Lod_num_bl[z][1]))
        # print("PO_num_li:"+str(PO_num_li))
        PO_num_li.append(PO_num) # 2-1-2-3-2

#print(Lod_num_bl) # None 인 모든 셀 정보 출력
#print(Lod_num) # None이 아닌 모든 셀 정보 출력
print("Load 넘버가 없는 모든 PO번호:"+ str(PO_num_li)) # Load 넘버가 없는 모든 PO번호 출력

# List index 값확인
# index = Lod_num_bl.index([9,4]) 
# print('The index of 9,4:', index)

#드라이버 로딩 
driver = webdriver.Chrome('./chromedriver.exe') 

#사용할 변수 선언 
#url = 'https://retaillink.login.wal-mart.com/?resumePath=/as/N3SUi/resume/as/authorization.ping' << 사용불가&로그인 불가 url 
#url = 'https://retaillink.login.wal-mart.com/?resumePath=/as/ie6cU/resume/as/authorization.ping' << 사용불가&로그인 불가 url
url = 'https://retaillink2.wal-mart.com/TSCP2/?ukey=W6520'
uid = 'chany.oh@cjlogisticsamerica.com' 
upw = 'CJL123!@#' 

#Walmart 로그인 페이지로 이동 
driver.get(url) 
time.sleep(10)  #로딩 대기


# find_element_by_id
# find_element_by_name
# find_element_by_xpath
# find_element_by_link_text
# find_element_by_partial_link_text
# find_element_by_tag_name
# find_element_by_class_name


# 아이디 입력폼 Rev-1 >> tag_id = driver.find_element_by_id('uname') 
# find_element_by_class_name 테그 안에 해당하는 문장을 인식하도록 설정
tag_id =driver.find_element_by_css_selector("input[data-automation-id='uname']")

#패스워드 입력폼 Rev-1 >> tag_pw = driver.find_element_by_id('pwd') 
tag_pw =driver.find_element_by_css_selector("input[data-automation-id='pwd']")


# id 입력 
# # 입력폼 클릭 -> paperclip에 선언한 uid 내용 복사 -> 붙여넣기 
tag_id.click() 
pyperclip.copy(uid) 
tag_id.send_keys(Keys.CONTROL, 'v') 
time.sleep(1) 

# pw 입력 # 입력폼 클릭 -> paperclip에 선언한 upw 내용 복사 -> 붙여넣기 
tag_pw.click() 
pyperclip.copy(upw) 
tag_pw.send_keys(Keys.CONTROL, 'v') 
time.sleep(1) 

#로그인 버튼 클릭 
login_btn = driver.find_element_by_css_selector("button[data-automation-id='loginBtn']")
login_btn.click() 
time.sleep(6) 

# Supplier 버튼 클릭
Supplier_btn = driver.find_element_by_id('transom-navbar-Supplier')
if Supplier_btn != None: # 2-1-4-2. [IF]: Transportation Supply Chain Portal 2.0 접속이 되었다면 
    Supplier_btn.click() 
else:
    Supply_Chain_btn = driver.find_element_by_link_text("Transportation Supply Chain Portal 2.0")
    Supply_Chain_btn.click() 
time.sleep(4) 

# Routing Tools 버튼 클릭
Routing_Tools_btn = driver.find_element_by_id('twist-nav-Routing_Tools')
Routing_Tools_btn.click() 
time.sleep(2) 

# Routing Status 버튼 클릭
# a tag link_text를 직접 클릭할 수 있는 모듈 사용함
Routing_Status_btn = driver.find_element_by_link_text("Routing Status")
Routing_Status_btn.click() 
time.sleep(2) 

# PO# or Load# 버튼 클릭
# PO_Load_btn = driver.find_element_by_css_selector("div[id='mat-tab-label-0-1']").click()
PO_Load_btn = driver.find_element_by_id("mat-tab-label-0-1")
PO_Load_btn.click() 
time.sleep(2) 

# Walmart PO번호 입력 
# localExcel.PO_num_li #PO 번호 리스트 Import
import localExcel
tag_id =driver.find_element_by_id("mat-input-0")
Po_bl_num = str(localExcel.PO_num_li)
tag_id.click() 
pyperclip.copy(Po_bl_num) 
tag_id.send_keys(Keys.CONTROL, 'v') 
#localExcel.PO_num_li=[] #초기화
time.sleep(2)


# 마우스 이동 (x 좌표, y 좌표)
pyautogui.moveTo(500, 500)

# 마우스 클릭
pyautogui.click()
time.sleep(2)

#SEARCH 버튼 클릭 
search_btn = driver.find_element_by_xpath(".//span[@class = 'mat-button-wrapper' and contains(text(), 'Search')]")
search_btn.click() 
time.sleep(2) 


#체크 박스 클릭 
login_btn = driver.find_element_by_css_selector("input[type='checkbox']")
login_btn.click() 
time.sleep(2) 

#Export 버튼 클릭 
export_btn = driver.find_element_by_xpath("//*[contains(text(), 'Export')]")
export_btn.click() 
time.sleep(5) 



for y in range(len(Lod_num_bl)): # 2-1-2
    Lod_num_bl[y][1]-=2 # 2-1-2-1
    
for z in range(len(Lod_num_bl)): # 2-1-2-3
    for k in range(0,1):
        PO_num=ws.cell(row=Lod_num_bl[z][k], column=Lod_num_bl[z][1]).value # 2-1-2-3-1
        # print("row:"+str(Lod_num_bl[z][k]))
        # print("column:"+str(Lod_num_bl[z][1]))
        # print("PO_num_li:"+str(PO_num_li))
        PO_num_li.append(PO_num) # 2-1-2-3-2

#print(Lod_num_bl) # None 인 모든 셀 정보 출력
#print(Lod_num) # None이 아닌 모든 셀 정보 출력
print("Load 넘버가 없는 모든 PO번호:"+ str(PO_num_li)) # Load 넘버가 없는 모든 PO번호 출력


## 현재 폴더에서 파일 이름 리스트를 얻기
path = "C:/Users/zinus_user/Downloads"
file_list = os.listdir(path)
print ("file_list: {}".format(file_list))

## 파일중 엑셀파일 리스트 얻기
exfile_list =[i for i in file_list if os.path.splitext(i)[-1]==".xlsx"]
print("excel:", exfile_list)

len_num=len(exfile_list)-1 # 항상 가장 마지막에 다운 받은 엑셀 파일기준으로 작업
excelfilename = exfile_list[len_num]
wb1 = load_workbook(str(path+"/"+excelfilename)) # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
ws1 = wb1.active  # 활성화된 Sheet
print("Latest excel file name", excelfilename)


k=2
y=4
l=2

for u in range(len(localExcel.PO_num_li)): 
    print("Po_num_li:"+str(localExcel.PO_num_li[u]))
    z=2
    for z in range (2, 150):
        # 2 부터 150번 까지 
        # "Walmart PIckups GA Warehouse.xlsx" 에서 Po번호 같은 번호 찾기
        verify=ws.cell(row=z, column=2).value # "Walmart PIckups GA Warehouse.xlsx" 에서 Po 번호 컬럼에서 리스트상에 있는 Po번호가 같은 번호를 식별
        print("verify:"+str(verify))
        if str(localExcel.PO_num_li[u]) == str(verify) : # 값이 같다면 
            index = ws1.cell(row=l, column=2).value #Routing_Status_11-17-2021_022456.xlsx의 Load 번호 다운 받은 PO 파일 
            print("index:"+str(index))
            
            ws.cell(row=z, column=y, value=str(index)) # "Walmart PIckups GA Warehouse.xlsx"의 빈 Load cell에 값을 넣어라
            log2=ws.cell(row=z, column=y, value=index)
            print("log2:"+str(log2))
            
            wb.save("Sheet1.csv")
            l+=1
        z+=1
    
