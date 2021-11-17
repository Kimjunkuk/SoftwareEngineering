from openpyxl import load_workbook # 파일 불러오기
import localExcel
import os


def insert_load_num():
    ## 현재 폴더에서 파일 이름 리스트를 얻기
    path = "C:/Users/zinus_user/Downloads"
    file_list = os.listdir(path)
    print ("file_list: {}".format(file_list))

    ## 파일중 엑셀파일 리스트 얻기
    exfile_list =[i for i in file_list if os.path.splitext(i)[-1]==".xlsx"]
    print("excel:", exfile_list)
    
    len_num=len(exfile_list)-1 # 항상 가장 마지막에 다운 받은 엑셀 파일기준으로 작업
    excelfilename = exfile_list[len_num]
    wb1 = load_workbook(str(path+"/"+excelfilename)) # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
    ws1 = wb1.active  # 활성화된 Sheet
    print("Latest excel file name", excelfilename)

    # 엑셀 시트 이름 얻기
    # wb = load_workbook(filename = "C:/Users/zinus_user/Downloads"+"/"+str(excelfilename))
    # print (wb.sheetnames)
    # ['1번', 'two']
    wb2 = load_workbook("Walmart PIckups GA Warehouse.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
    ws2 = wb2.active  # 활성화된 Sheet
    
    #반복
        # Load 번호가 없는 PO번호의 Load번호 셀 확인 
        # 값 삽입   
    k=2
    y=4
    z=2 
    
    for u in range(len(localExcel.PO_num_li)): 
        for z in range (2, 150):
        # 2 부터 150번 까지 
        # "Walmart PIckups GA Warehouse.xlsx" 에서 Po번호 같은 번호 찾기
            verify=ws2.cell(row=z, column=2).value # Po 번호 컬럼에서 리스트상에 있는 Po번호가 같은 번호를 식별
            if str(localExcel.PO_num_li[u]) == verify : # 값이 같다면 
                index = ws1.cell(row=z, column=2).value
                ws2.cell(row=k, column=y, value=index) # Load cell에 값을 넣어라
                wb2.save("Walmart PIckups GA Warehouse.xlsx")
        


insert_load_num()

# 성공 코드 
wb = load_workbook("Walmart PIckups GA Warehouse.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
ws = wb.active  # 활성화된 Sheet

wb1 = load_workbook("C:/Users/zinus_user/Downloads/Routing_Status_11-17-2021_022456.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
ws1 = wb1.active  # 활성화된 Sheet

k=2
y=4
l=2

for u in range(len(localExcel.PO_num_li)): 
    print("Po_num_li:"+str(localExcel.PO_num_li[u]))
    z=2
    for z in range (2, 150):
        # 2 부터 150번 까지 
        # "Walmart PIckups GA Warehouse.xlsx" 에서 Po번호 같은 번호 찾기
        verify=ws.cell(row=z, column=2).value # "Walmart PIckups GA Warehouse.xlsx" 에서 Po 번호 컬럼에서 리스트상에 있는 Po번호가 같은 번호를 식별
        print("verify:"+str(verify))
        if str(localExcel.PO_num_li[u]) == str(verify) : # 값이 같다면 
            index = ws1.cell(row=l, column=2).value #Routing_Status_11-17-2021_022456.xlsx의 Load 번호 다운 받은 PO 파일 
            print("index:"+str(index))
            
            ws.cell(row=z, column=y, value=str(index)) # "Walmart PIckups GA Warehouse.xlsx"의 빈 Load cell에 값을 넣어라
            log2=ws.cell(row=z, column=y, value=index)
            print("log2:"+str(log2))
            
            wb.save("Walmart PIckups GA Warehouse.xlsx")
            l+=1
        z+=1
