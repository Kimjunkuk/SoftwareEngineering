from oauth2client.client import verify_id_token
from openpyxl import load_workbook # 파일 불러오기
from bs4 import BeautifulSoup
from selenium import webdriver
import time

def localExcel():
    wb = load_workbook("Walmart PIckups GA Warehouse.xlsx") # Walmart PIckups GA Warehouse.xlsx 파일에서 wb 을 불러옴
    ws = wb.active  # 활성화된 Sheet

    """
    1. Google sheet download (Walmart PIckups GA Warehouse.xlsx) 

    2. 빈 Load 번호 cell 정보 확인 & 빈 Load번호에 해당하는 PO 번호 식별하여 값 조회작업
        2-1. [IF]: 빈칸이 있다면
            2-1-1. 빈 Load 번호 cell 정보 확인
                2-1-1-1. Blank가 아닌 PO번호의 셀 정보를 담을 리스트 생성
                2-1-1-2. Blank인 PO번호의 셀 정보를 담을 리스트 생성
                2-1-1-3. [Loop]: 2번 row 부터 ws.max_row 마지막 row 까지 작업 
                    2-1-1-3-1. y변수와 컬럼의 수를 넣어 식별한 컬럼을 지정하여 준다. 
                    2-1-1-3-2. Load 넘버 컬럼의 값을 마지막 row 까지 가져와 변수에 넣는다.
                    2-1-1-3-3. [IF]: Load 컬럼의 cell이 None 이라면?
                        2-1-1-3-3-1. Load 컬럼의 cell이 None인 셀 정보를 중첩 리스트형태로 Lod_num_bl 리스트에 넣는다. 
                    
                    2-1-1-3-4. [IF]: Load 컬럼의 cell이 None 아니라면?
                        2-1-1-3-4-1. Load 컬럼의 cell이 None인 셀 정보를 중첩 리스트형태로 Lod_num 리스트에 넣는다. 
                
            2-1-2. [Loop]: 리스트의 길이만큼 [][1] << 두번째 인덱스의 모든 값을 불러온다  
                2-1-2-1. 뒷자리에 해당 하는 값을 모두 -2 씩 감소 시켜 다시 Lod_num_bl 중첩 리스트에 저장한다.
                2-1-2-2. PO번호를 담을 전역 리스트 추가 생성
                2-1-2-3. [Loop]: Lod_num_bl 각각의 인덱스에 해당하는 셀정보(PO번호)를 저장된 리스트의 길이만큼 반복 식별
                    2-1-2-3-1. 변수를 생성하여 Load 번호가 없는 PO번호를 저장함
                    2-1-2-3-2. 변수에 저장된 PO 번호를 미리 생성한 전역 리스트에 저장 
                


            2-1-4. Walmart supplier 웹사이트에 접속
                2-1-4-1. [IF]: Ratail link 웹사이트로 접속이 되었다면 
                    2-1-4-1-1. Transportation Supply Chain Portal 2.0 메뉴 클릭
                    2-1-4-1-2. 2-1-4-2-1 << 부터 작업시작
                    
                2-1-4-2. [IF]: Transportation Supply Chain Portal 2.0 접속이 되었다면 
                    2-1-4-2-1. ID, PW 정보 입력 로그인 버튼 클릭
                    2-1-4-2-2. 상단 Supplier 메뉴 클릭
                    2-1-4-2-3. Routing Tools 메뉴 클릭
                    2-1-4-2-4. Routing Status 메뉴 클릭
                    2-1-4-2-5. Search 메뉴에서 PO# or Load# 메뉴 클릭
                    2-1-4-2-6. Search By PO Numbers 메뉴에 Load 번호가 없는 PO 번호 입력
                    2-1-4-2-7. 하단의 Search 메뉴 클릭 
                    

            
        2-2. [IF]: 빈칸이 없다면 
            2-2-1. 다음 확인 작업시 까지 시스템 대기
        
    3. 

    """


    Lod_num = []  # 2-1-1-1
    Lod_num_bl = [] # 2-1-1-2


    for x in range(2, 150):# 2-1-1-3
        y=4 # 2-1-1-3-1
        verify=ws.cell(row=x, column=y).value # 2-1-1-3-2
        if verify == None: #
            # 2-1-1-3-3-1
            #Lod_num_bl.append(ws.cell(row=x, column=y)) 
            Lod_num_bl.append([x,y]) # << Nested Lists(중첩 리스트 생성 )
    

        else: # 2-1-1-3-4
            #Lod_num.append(ws.cell(row=x, column=y)) 
            Lod_num.append([x,y]) # 2-1-1-3-4-1


    # List index 값확인
    # index = Lod_num_bl.index([9,4]) 
    # print('The index of 9,4:', index)

    PO_num_li = []

    for y in range(len(Lod_num_bl)): # 2-1-2
        Lod_num_bl[y][1]-=2 # 2-1-2-1
        
    for z in range(len(Lod_num_bl)): # 2-1-2-3
        for k in range(0,1):
            PO_num=ws.cell(row=Lod_num_bl[z][k], column=Lod_num_bl[z][1]).value # 2-1-2-3-1
            # print("row:"+str(Lod_num_bl[z][k]))
            # print("column:"+str(Lod_num_bl[z][1]))
            # print("PO_num_li:"+str(PO_num_li))
            PO_num_li.append(PO_num) # 2-1-2-3-2

    #print(Lod_num_bl) # None 인 모든 셀 정보 출력
    #print(Lod_num) # None이 아닌 모든 셀 정보 출력
    #print(PO_num_li) # Load 넘버가 없는 모든 PO번호 출력


localExcel()
