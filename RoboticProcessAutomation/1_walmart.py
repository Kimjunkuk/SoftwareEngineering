from oauth2client.client import verify_id_token
from openpyxl import load_workbook # 파일 불러오기
from bs4 import BeautifulSoup
from selenium import webdriver
import time

wb = load_workbook("Walmart PIckups GA Warehouse.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

"""

D컬럼(Load#) 공백인 PO#를 식별 

만약 D컬럼의 값이 공백일 경우 mmm

"""    

# cell 데이터 불러오기
# Blank가 아닌 PO#의 셀, Load 넘버가 빈 PO#의 넘버 

Po_num = [] # Blank가 아닌 PO번호의 셀 정보를 담을 리스트 생성 
Po_num_bl = [] # Blank인 PO번호의 셀 정보를 담을 리스트 생성

# a=[]
# for i in range(5):    
#     a.append(i)
# print(a)
#print(ws.cell(row=6, column=2).value)
for x in range(2, 149):
    y=2
    verify=ws.cell(row=x, column=y).value #PO 넘버 컬럼의 값을 149low 까지 가져와 변수에 넣는다.
    if verify == None: #PO#가 null 아닐 경우
        #PO#의 셀 정보를 리스트에 넣는다. 
        Po_num.append(ws.cell(row=x, column=y))  #리스트 2번 부터 데이터 삽입 #셀의 정보
 

    elif verify != "": #PO번호가 null 일경우 
        Po_num_bl.append(ws.cell(row=x, column=y)) #Po_num_bl 리스트에 PO번호가 null인 셀의 정보를 순차 삽입
        #print(Po_num_bl) # 모든 Po_num_bl 리스트의 값을 출력 

print(Po_num)


    #print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
    # verify=ws.cell(row=x, column=y).value
    #if verify=="": #<< 공백일 경우에만 작용

# driver = webdriver.Chrome("C:\Program Files\Google\Chrome\Application\chrome_proxy.exe")
# url = 'https://retaillink.login.wal-mart.com/?resumePath=/as/N3SUi/resume/as/authorization.ping'
# driver.get(url)
#delay=3 #리소스를 기다리는 시간 
#driver.implicitly_wait(delay)





# driver.find_element_by_id('uname').send_keys('chany.oh@cjlogisticsamerica.com')
# driver.find_element_by_id('pwd').send_keys('CJL123!@#')
# driver.find_element_by_id('loginBtn').click()

# find_element_by_id
# find_element_by_name
# find_element_by_xpath
# find_element_by_link_text
# find_element_by_partial_link_text
# find_element_by_tag_name
# find_element_by_class_name
# find_element_by_css_selector

        


#for x in range(5):
    """[summary]
    """ #   print(ws["A"+x].value)
