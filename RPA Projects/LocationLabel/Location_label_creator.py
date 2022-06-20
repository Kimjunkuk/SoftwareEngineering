## [ Library list in use (Korean version)]
import qrcode
from openpyxl import Workbook
from types import NoneType
import pygsheets  # 1
from openpyxl import load_workbook  # 2
import os  # 8
from tkinter import *
from tkinter import *
import os

from MergePDF import merge_all_pdf_class
cj=merge_all_pdf_class()



"""
한빈 데이터 입력 작업 
1. 한빈의 시작 단수와 최대 단수는 어떻게 되는지 파악 (완료)
2. 시작 단수와 최대 단수를 만족하는 시트에 로케이션 정보 입력 (완료)
3. 로케이션 정보에 해당하는 QR 코드 및 데이터 삽입 (완료)
4. 작업을 진행한 시트의 페이지를 PDF로 출력 저장(완료)
5. 현재까지 출력된 모든 PDF 순서대로 병합 (완료)


"""



# storage_bin = data_location.cell(row=x, column=y).value
# next_storage_bin = data_location.cell(row=z, column=y).value

# tunnel_bin_check = storage_bin[5:7]  # 빈 번호
# level_verify = storage_bin[8:9]  # level




x = 1  # now row
y = 1  # now column
z = 2  # next row
page_num=1
qrcode_num=1

while True:

    # 데이터 정재 작업을 수행할 엑셀 시트 불러옴
    wb = load_workbook("./LocationLabel/Location Label_012122.xlsx")

    # 불러온 엑셀 시트 활성화
    ws = wb.active  

    # 엑셀 파일 시트 선택
    data_location = wb["Data"]

    try:

        bin_l = []  # Left side
        bin_m = []  # Middle side
        bin_r = []  # Right side
        
        storage_bin = data_location.cell(row=x, column=y).value
        if storage_bin == None or storage_bin == NoneType:
            print("End process now_bin: [" +str(storage_bin) +"], next_bin:"+str(next_storage_bin))
            if str(side_verify) == "L":
                print("now_bin: ["+str(storage_bin)+"], L")
                bin_l.append(storage_bin)

            elif str(side_verify) == "R":
                print("now_bin: ["+str(storage_bin)+"], R")
                bin_r.append(storage_bin)

            else:
                print("now_bin: ["+str(storage_bin)+"], M")
                bin_m.append(storage_bin)

            print("End process")
            break
        # 리스트에 적재하여 같은 빈의 최대 적재 가능 층을 파악, 같은 빈의 위치 정보를 방향 별로 분류함

        while True:
            # 현재 빈 번호와 다음 빈 번호가 같다면 다중 리스트에 값 적재


            storage_bin = data_location.cell(row=x, column=y).value
            next_storage_bin = data_location.cell(row=z, column=y).value
            side_verify = storage_bin[9:10]  # 번호의 방향을 확인 한다

            try: # now_bin, next_bin 의 NoneType 데이터가 유입되는 것을 무시하도록 예외처리함
                now_bin = storage_bin[5:7]  # 빈 번호
                next_bin = next_storage_bin[5:7]  # 빈 번호
            except Exception as e:
                break
                print(e)

            if int(now_bin) == int(next_bin):

                print("now_bin: [" +str(storage_bin) +"], next_bin:"+str(next_storage_bin))
                if str(side_verify) == "L":
                    print("now_bin: ["+str(storage_bin)+"], L")
                    bin_l.append(storage_bin)
                    x+=1
                    z+=1

                elif str(side_verify) == "R":
                    print("now_bin: ["+str(storage_bin)+"], R")
                    bin_r.append(storage_bin)
                    x+=1
                    z+=1

                else:
                    print("now_bin: ["+str(storage_bin)+"], M")
                    bin_m.append(storage_bin)
                    x+=1
                    z+=1



            # 현재 빈 번호와 다음 빈 번호가 같지 않으면 현재 빈 번호를 방향에 맞게 리스트에 적재하고 연산 종료
            elif int(now_bin) != next_bin:

                print("elif >> now_bin: [" +str(storage_bin) +"], next_bin:"+str(next_storage_bin))
                if str(side_verify) == "L":
                    print("now_bin: ["+str(storage_bin)+"], L")
                    bin_l.append(storage_bin)

                elif str(side_verify) == "R":
                    print("now_bin: ["+str(storage_bin)+"], R")
                    bin_r.append(storage_bin)

                else:
                    print("now_bin: ["+str(storage_bin)+"], M")
                    bin_m.append(storage_bin)
                break


        # 왼쪽 또는 오른쪽에 적재된 마지막 값의 층수 확인를 확인하여 해당 빈은 몇 단까지 존재하는지 파악
        last_element = bin_l[-1]
        first_element = bin_l[0]
        storage_bin_num = bin_l[0]
        print("last_element:"+str(last_element)+", first_element:"+str(first_element))

        max_lv = last_element[8:9]  # max level
        min_lv = first_element[8:9] # min level 
        bin_num = storage_bin_num[5:7] # bin num
        print("max_lv:"+str(max_lv)+", min_lv:"+str(min_lv)+", bin_num:"+str(bin_num))

        print("left bin: " + str(bin_l))
        print("middle bin: " + str(bin_m))
        print("right bin: " + str(bin_r))


        while True:
            print("start here")

            if str(bin_num) == str("27") or str(bin_num) == str("28"): # Tunnel bin 

                print("Tunnel bin")

                if str(min_lv)=="3" and str(max_lv)=="4":

                    from openpyxl.drawing.image import Image

                    tunnelLv4_location = wb["TunnelLv4"] # 빈의 최소 시작 높이1, 최대 높이 4 임으로 Lv4 시트 선택

                    # 오른쪽, 왼쪽 화살표 이미지 불러오기
                    arrow_right_img = Image("./LocationLabel/right_arrow.png")
                    arrow_left_img = Image("./LocationLabel/left_arrow.png")

                    # 오른쪽 화살표 사이즈 수정
                    arrow_right_img.height = 95
                    arrow_right_img.width = 282

                    # 왼쪽 화살표 사이즈 수정
                    arrow_left_img.height = 95
                    arrow_left_img.width = 282

                    tunnelLv4_location.add_image(arrow_right_img, 'A1') 
                    tunnelLv4_location.add_image(arrow_left_img, 'I1') 

                    if len(bin_l) >= 1: # 왼쪽 값이 존재 한다면 

                        row_num=9
                        row_num_for_bin_data=10

                        for i in range(len(bin_l)):
                            
                            tunnelLv4_location.cell(row=row_num_for_bin_data, column=1, value=bin_l[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_l[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv4_location.add_image(img, 'D'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                            

                    if len(bin_r) >= 1:

                        row_num=9
                        row_num_for_bin_data=10

                        for i in range(len(bin_r)):
                            
                            tunnelLv4_location.cell(row=row_num_for_bin_data, column=9, value=bin_r[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_r[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv4_location.add_image(img, 'L'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                        
                        

                    if len(bin_m) >= 1:

                        row_num=9
                        row_num_for_bin_data=10

                        for i in range(len(bin_m)):
                            
                            tunnelLv4_location.cell(row=row_num_for_bin_data, column=5, value=bin_m[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_m[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv4_location.add_image(img, 'H'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                                


                    wb.save("./LocationLabel/Location Label_012122.xlsx")

                    #pdf로 변환할 파일명 선택 
                    wb = excel.Workbooks.Open("C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/Location Label_012122.xlsx") 

                    #워크북의 시트명 설정 
                    ws_sht = wb.Worksheets("TunnelLv4") 

                    #설정한 시트 선택 
                    ws_sht.Select() 

                    #PDF파일을 저장할 경로 및 파일명 지정 
                    savepath = "C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/pdf/sample"+str(page_num)+"_pdf.pdf"

                    #활성화된 시트를 pdf 저장 
                    wb.ActiveSheet.ExportAsFixedFormat(0, savepath) 

                    


                if str(min_lv)=="3" and str(max_lv)=="5":

                    from openpyxl.drawing.image import Image

                    tunnelLv5_location = wb["TunnelLv5"] # 빈의 최소 시작 높이1, 최대 높이 4 임으로 Lv4 시트 선택

                    # 오른쪽, 왼쪽 화살표 이미지 불러오기
                    arrow_right_img = Image("./LocationLabel/right_arrow.png")
                    arrow_left_img = Image("./LocationLabel/left_arrow.png")

                    # 오른쪽 화살표 사이즈 수정
                    arrow_right_img.height = 95
                    arrow_right_img.width = 282

                    # 왼쪽 화살표 사이즈 수정
                    arrow_left_img.height = 95
                    arrow_left_img.width = 282

                    tunnelLv5_location.add_image(arrow_right_img, 'A1') 
                    tunnelLv5_location.add_image(arrow_left_img, 'I1') 


                    if len(bin_l) >= 1: # 왼쪽 값이 존재 한다면 

                        row_num=13
                        row_num_for_bin_data=14

                        for i in range(len(bin_l)):
                            
                            tunnelLv5_location.cell(row=row_num_for_bin_data, column=1, value=bin_l[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_l[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv5_location.add_image(img, 'D'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                            

                    if len(bin_r) >= 1:

                        row_num=13
                        row_num_for_bin_data=14

                        for i in range(len(bin_r)):
                            
                            tunnelLv5_location.cell(row=row_num_for_bin_data, column=9, value=bin_r[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_r[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv5_location.add_image(img, 'L'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                            
                            

                    if len(bin_m) >= 1:

                        row_num=13
                        row_num_for_bin_data=14

                        for i in range(len(bin_m)):
                            
                            tunnelLv5_location.cell(row=row_num_for_bin_data, column=5, value=bin_m[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_m[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv5_location.add_image(img, 'H'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                                    


                    wb.save("./LocationLabel/Location Label_012122.xlsx")

                    #pdf로 변환할 파일명 선택 
                    wb = excel.Workbooks.Open("C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/Location Label_012122.xlsx") 

                    #워크북의 시트명 설정 
                    ws_sht = wb.Worksheets("TunnelLv5") 

                    #설정한 시트 선택 
                    ws_sht.Select() 

                    #PDF파일을 저장할 경로 및 파일명 지정 
                    savepath = "C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/pdf/sample"+str(page_num)+"_pdf.pdf"

                    #활성화된 시트를 pdf 저장 
                    wb.ActiveSheet.ExportAsFixedFormat(0, savepath) 




                if str(min_lv)=="3" and str(max_lv)=="6":

                    from openpyxl.drawing.image import Image

                    tunnelLv6_location = wb["TunnelLv6"] # 빈의 최소 시작 높이1, 최대 높이 4 임으로 Lv4 시트 선택

                    # 오른쪽, 왼쪽 화살표 이미지 불러오기
                    arrow_right_img = Image("./LocationLabel/right_arrow.png")
                    arrow_left_img = Image("./LocationLabel/left_arrow.png")

                    # 오른쪽 화살표 사이즈 수정
                    arrow_right_img.height = 95
                    arrow_right_img.width = 282

                    # 왼쪽 화살표 사이즈 수정
                    arrow_left_img.height = 95
                    arrow_left_img.width = 282

                    tunnelLv6_location.add_image(arrow_right_img, 'A1') 
                    tunnelLv6_location.add_image(arrow_left_img, 'I1') 

                    if len(bin_l) >= 1: # 왼쪽 값이 존재 한다면 

                        row_num=17
                        row_num_for_bin_data=18

                        for i in range(len(bin_l)):
                            
                            tunnelLv6_location.cell(row=row_num_for_bin_data, column=1, value=bin_l[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_l[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv6_location.add_image(img, 'D'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                            

                    if len(bin_r) >= 1:

                        row_num=17
                        row_num_for_bin_data=18

                        for i in range(len(bin_r)):
                            
                            tunnelLv6_location.cell(row=row_num_for_bin_data, column=9, value=bin_r[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_r[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv6_location.add_image(img, 'L'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                        
                        

                    if len(bin_m) >= 1:

                        row_num=17
                        row_num_for_bin_data=18

                        for i in range(len(bin_m)):
                            
                            tunnelLv6_location.cell(row=row_num_for_bin_data, column=5, value=bin_m[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_m[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            tunnelLv6_location.add_image(img, 'H'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                                


                    wb.save("./LocationLabel/Location Label_012122.xlsx")

                    #pdf로 변환할 파일명 선택 
                    wb = excel.Workbooks.Open("C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/Location Label_012122.xlsx") 

                    #워크북의 시트명 설정 
                    ws_sht = wb.Worksheets("TunnelLv6") 

                    #설정한 시트 선택 
                    ws_sht.Select() 

                    #PDF파일을 저장할 경로 및 파일명 지정 
                    savepath = "C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/pdf/sample"+str(page_num)+"_pdf.pdf"

                    #활성화된 시트를 pdf 저장 
                    wb.ActiveSheet.ExportAsFixedFormat(0, savepath) 


            

            ######## not tunnel bin ########
            

            #win32com.client 모듈 임포트 
            import win32com.client 

            #엑셀을 실행할 객체 생성 
            excel = win32com.client.Dispatch("Excel.Application") 

            if str(bin_num) != "27" or str(bin_num) != "28":

                print("Not tunnel bin ")
                    
                if str(min_lv)=="1" and str(max_lv)=="4":

                    from openpyxl.drawing.image import Image

                    Lv4_location = wb["Lv4"] # 빈의 최소 시작 높이1, 최대 높이 4 임으로 Lv4 시트 선택

                    # 오른쪽, 왼쪽 화살표 이미지 불러오기
                    arrow_right_img = Image("./LocationLabel/right_arrow.png")
                    arrow_left_img = Image("./LocationLabel/left_arrow.png")

                    # 오른쪽 화살표 사이즈 수정
                    arrow_right_img.height = 95
                    arrow_right_img.width = 282

                    # 왼쪽 화살표 사이즈 수정
                    arrow_left_img.height = 95
                    arrow_left_img.width = 282

                    Lv4_location.add_image(arrow_right_img, 'A1') 
                    Lv4_location.add_image(arrow_left_img, 'I1') 

                    if len(bin_l) >= 1: # 왼쪽 값이 존재 한다면 

                        row_num=17
                        row_num_for_bin_data=18
                        try:
                            for i in range(len(bin_l)):
                                Lv4_location.cell(row=row_num_for_bin_data, column=1, value=bin_l[i])
                                
                                # QR 코드 생성
                                img = qrcode.make(str(bin_l[i]))
                                type(img)  # qrcode.image.pil.PilImage
                                img.save("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                                # 이미지 불러오기
                                # time.sleep(10)
                                # print("dely 10")
                                img = Image("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                                # QR 코드 이미지 사이즈 수정
                                img.height = 100
                                img.width = 100

                                Lv4_location.add_image(img, 'D'+str(row_num))

                                row_num-=4
                                row_num_for_bin_data-=4
                                qrcode_num+=1
                                
                        except Exception as e:
                            print(e)
                            

                    if len(bin_r) >= 1:

                        row_num=17
                        row_num_for_bin_data=18
                        try:
                            for i in range(len(bin_r)):
                                
                                Lv4_location.cell(row=row_num_for_bin_data, column=9, value=bin_r[i])

                                # QR 코드 생성
                                img = qrcode.make(str(bin_r[i]))
                                type(img)  # qrcode.image.pil.PilImage
                                img.save("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                                # 이미지 불러오기
                                # time.sleep(10)
                                # print("dely 10")
                                img = Image("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                                # QR 코드 이미지 사이즈 수정
                                img.height = 100
                                img.width = 100

                                Lv4_location.add_image(img, 'L'+str(row_num))

                                row_num-=4
                                row_num_for_bin_data-=4
                                qrcode_num+=1
                                
                        except Exception as e:
                            print(e)
                        
                        

                    if len(bin_m) >= 1:

                        row_num=17
                        row_num_for_bin_data=18
                        try:
                            for i in range(len(bin_m)):
                                                                
                                Lv4_location.cell(row=row_num_for_bin_data, column=5, value=bin_m[i])

                                # QR 코드 생성
                                img = qrcode.make(str(bin_m[i]))
                                type(img)  # qrcode.image.pil.PilImage
                                img.save("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                                # 이미지 불러오기
                                # time.sleep(10)
                                # print("dely 10")
                                img = Image("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                                # QR 코드 이미지 사이즈 수정
                                img.height = 100
                                img.width = 100

                                Lv4_location.add_image(img, 'H'+str(row_num))

                                row_num-=4
                                row_num_for_bin_data-=4
                                qrcode_num+=1
                                
                        except Exception as e:
                            print(e)        


                    wb.save("./LocationLabel/Location Label_012122.xlsx")

                    #pdf로 변환할 파일명 선택 
                    wb = excel.Workbooks.Open("C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/Location Label_012122.xlsx") 

                    #워크북의 시트명 설정 
                    ws_sht = wb.Worksheets("Lv4") 

                    #설정한 시트 선택 
                    ws_sht.Select() 

                    #PDF파일을 저장할 경로 및 파일명 지정 
                    savepath = "C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/pdf/sample"+str(page_num)+"_pdf.pdf"

                    #활성화된 시트를 pdf 저장 
                    wb.ActiveSheet.ExportAsFixedFormat(0, savepath) 

                if str(min_lv)=="1" and str(max_lv)=="5":


                    Lv5_location = wb["Lv5"]

                    # 오른쪽, 왼쪽 화살표 이미지 불러오기
                    arrow_right_img = Image("./LocationLabel/right_arrow.png")
                    arrow_left_img = Image("./LocationLabel/left_arrow.png")

                    # 오른쪽 화살표 사이즈 수정
                    arrow_right_img.height = 95
                    arrow_right_img.width = 282

                    # 왼쪽 화살표 사이즈 수정
                    arrow_left_img.height = 95
                    arrow_left_img.width = 282

                    Lv5_location.add_image(arrow_right_img, 'A1') 
                    Lv5_location.add_image(arrow_left_img, 'I1') 

                    if len(bin_l) >= 1: # 왼쪽 값이 존재 한다면 

                        row_num=21
                        row_num_for_bin_data=22

                        for i in range(len(bin_l)):
                            
                            Lv5_location.cell(row=row_num_for_bin_data, column=1, value=bin_l[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_l[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            Lv5_location.add_image(img, 'D'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                            

                    if len(bin_r) >= 1:

                        row_num=21
                        row_num_for_bin_data=22

                        for i in range(len(bin_r)):
                            
                            Lv5_location.cell(row=row_num_for_bin_data, column=9, value=bin_r[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_r[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            Lv5_location.add_image(img, 'L'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                        
                        

                    if len(bin_m) >= 1:

                        row_num=21
                        row_num_for_bin_data=22

                        for i in range(len(bin_m)):
                            
                            Lv5_location.cell(row=row_num_for_bin_data, column=5, value=bin_m[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_m[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            Lv5_location.add_image(img, 'H'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                                


                    wb.save("./LocationLabel/Location Label_012122.xlsx")
                    
                    #pdf로 변환할 파일명 선택 
                    wb = excel.Workbooks.Open("C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/Location Label_012122.xlsx") 

                    #워크북의 시트명 설정 
                    ws_sht = wb.Worksheets("Lv5") 

                    #설정한 시트 선택 
                    ws_sht.Select() 

                    #PDF파일을 저장할 경로 및 파일명 지정 
                    savepath = "C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/pdf/sample"+str(page_num)+"_pdf.pdf"

                    #활성화된 시트를 pdf 저장 
                    wb.ActiveSheet.ExportAsFixedFormat(0, savepath) 


                    

                if str(min_lv)=="1" and str(max_lv)=="6":


                    Lv6_location = wb["Lv6"]
                    # 오른쪽, 왼쪽 화살표 이미지 불러오기
                    arrow_right_img = Image("./LocationLabel/right_arrow.png")
                    arrow_left_img = Image("./LocationLabel/left_arrow.png")

                    # 오른쪽 화살표 사이즈 수정
                    arrow_right_img.height = 95
                    arrow_right_img.width = 282

                    # 왼쪽 화살표 사이즈 수정
                    arrow_left_img.height = 95
                    arrow_left_img.width = 282

                    Lv6_location.add_image(arrow_right_img, 'A1') 
                    Lv6_location.add_image(arrow_left_img, 'I1') 

                    if len(bin_l) >= 1: # 왼쪽 값이 존재 한다면 

                        row_num=25
                        row_num_for_bin_data=26

                        for i in range(len(bin_l)):
                            
                            Lv6_location.cell(row=row_num_for_bin_data, column=1, value=bin_l[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_l[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_l[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            Lv6_location.add_image(img, 'D'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                            

                    if len(bin_r) >= 1:

                        row_num=25
                        row_num_for_bin_data=26

                        for i in range(len(bin_r)):
                            
                            Lv6_location.cell(row=row_num_for_bin_data, column=9, value=bin_r[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_r[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_r[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            Lv6_location.add_image(img, 'L'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                        
                        

                    if len(bin_m) >= 1:

                        row_num=25
                        row_num_for_bin_data=26

                        for i in range(len(bin_m)):
                            
                            Lv6_location.cell(row=row_num_for_bin_data, column=5, value=bin_m[i])

                            # QR 코드 생성
                            img = qrcode.make(str(bin_m[i]))
                            type(img)  # qrcode.image.pil.PilImage
                            img.save("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # 이미지 불러오기
                            # time.sleep(10)
                            # print("dely 10")
                            img = Image("./LocationLabel/qrcode/"+str(bin_m[i])+".png")

                            # QR 코드 이미지 사이즈 수정
                            img.height = 100
                            img.width = 100

                            Lv6_location.add_image(img, 'H'+str(row_num))

                            # os.remove("./LocationLabel/qr_code.png")
                            row_num-=4
                            row_num_for_bin_data-=4
                                


                    wb.save("./LocationLabel/Location Label_012122.xlsx")
                    
                    #pdf로 변환할 파일명 선택 
                    wb = excel.Workbooks.Open("C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/Location Label_012122.xlsx") 

                    #워크북의 시트명 설정 
                    ws_sht = wb.Worksheets("Lv6") 

                    #설정한 시트 선택 
                    ws_sht.Select() 

                    #PDF파일을 저장할 경로 및 파일명 지정 
                    savepath = "C:/Users/Zinus_user/Desktop/Zinus_RPA_project/LocationLabel/pdf/sample"+str(page_num)+"_pdf.pdf"

                    #활성화된 시트를 pdf 저장 
                    wb.ActiveSheet.ExportAsFixedFormat(0, savepath) 



                

            # 엑셀 워크북 및 프로그램 종료 
            # # 종료를 제대로 해주어야 다음 실행시 에러 안생김 
            # wb.close()
            excel.Quit()
            page_num+=1
            x+=1 # 한빈의 연산이 끝난 후 다음 빈의 연산을 시작할 수 있도록 한다. 
            z+=1 # 다음빈의 다음 빈 번호를 식별할 수 있도록 한다. 

            from openpyxl import load_workbook  # 2
            wb = load_workbook("./LocationLabel/Location Label_012122.xlsx")
            # wb['Lv4']._images
            # del wb['Lv4']._images [11]
            print("delete start")
            images = []
            for sheet in wb:     
                for image in sheet._images:
                    if image.path[-3:] !="png":
                        images.append(image)
                        
                sheet._images =   images
            sheet._images = []
            wb.save("./LocationLabel/Location Label_012122.xlsx")
            print("delete end")
            break    
        
        
        
    except Exception as e:
        print(e)
        # wb.Close(False)
        excel.Quit()
        break

cj.merge_all_pdf()







